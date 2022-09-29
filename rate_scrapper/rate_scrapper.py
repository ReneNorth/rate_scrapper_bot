from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import pandas as pd
import sqlite3


# links
link_first_part = 'https://nationalbank.kz/ru/exchangerates/ezhednevnye-oficialnye-rynochnye-kursy-valyut/excel?'
test_link = 'https://nationalbank.kz/ru/exchangerates/ezhednevnye-oficialnye-rynochnye-kursy-valyut/excel?rates%5B%5D=5&rates%5B%5D=6&rates%5B%5D=16&rates%5B%5D=23&beginDate=2022-09-01&endDate=2022-09-27'


# READ EXCEL
#wb = load_workbook(filename='rates.xlsx', read_only=True)
excel_file = 'rates.xlsx'
rates = pd.read_excel(excel_file)



# SQLITE https://docs.python.org/3/library/sqlite3.html
# SQLITE + PANDAS

# First, we need to create a new database and open a database connection to allow sqlite3 to work with it.
con = sqlite3.connect("rates_db.db")

# In order to execute SQL statements and fetch results from SQL queries, we will need to use a database cursor
cur = con.cursor()





# TEMP
id_rates = [5, 6, 16, 23]
begin_date = '2022-01-01'
end_date = '2022-09-27'

# что тестировать 
# ответы от API 
# соответствие формата файла
# путь к файлу
# как правильно поступать с файлом после добавления в DB? Удалять?
# отдельно протестировать добавление новых строк
# добавить try/exception везде



def form_link(id_rates, begin_date, end_date):
    """ Fuction for creating a link based on the currency input and dates """
    mystring = 'rates%5B%5D='

    currencies_list = [mystring + id + '&' for id in map(str, id_rates)]
    currencies = ''.join(currencies_list)

    return f'{link_first_part}{currencies}beginDate={begin_date}&endDate={end_date}'


# https://www.dataquest.io/blog/excel-and-pandas/
rates_to_pands = pd.ExcelFile(excel_file)
rates_list = []
rates_list.append(rates_to_pands.parse('Courses'))
rates = pd.concat(rates_list)
rates = rates.set_index('Date')
#rates.rename(columns=rates.iloc[1])
print(rates.head(10))
# print(rates.iloc[1]['USD'])


date_01_usd = '06.01.2022'
currency_1 = 'USD'

date_02_eur = '06.01.2022'
currency_2 = 'RUB'

print(f'Курс {currency_1} на {date_01_usd}:', rates.loc[f'{date_01_usd}', f'{currency_1}'])
print(f'Курс {currency_2} на {date_02_eur}:', rates.loc[f'{date_02_eur}', f'{currency_2}'])


# Write the dataframe to SQLite DB  'rates_db'
rates.to_sql('rates_db', con, if_exists="replace")
# or replace  https://pandas.pydata.org/docs/reference/api/pandas.DataFrame.to_sql.html

# con.close()
print('SQLite below')
print('---')
print('---')

def rate_on_date():
    list_of_currencies = ', '.join([currency_1, currency_2])
    sql_query = f'SELECT {list_of_currencies} FROM rates_db WHERE Date="{date_01_usd}"'

    cur.execute(f'SELECT {currency_1}, {currency_2} FROM rates_db WHERE Date="{date_01_usd}"')
    res = cur.fetchall()
    return res

# print(res)

# print(f'Курс {currency_1} на {date_01_usd}:', rates.loc[f'{date_01_usd}', f'{currency_1}'])
# print(f'Курс {currency_2} на {date_02_eur}:', rates.loc[f'{date_02_eur}', f'{currency_2}'])


def main():
    pass


if __name__ == '__main__':
    main()





# print(rates.iloc[7, 2])




#courses_wb = wb['Courses']
#print(courses_wb['A1'].value)

# пробуем извлечь курс доллара 01.01.2022	1	431,8




# print(form_link(id_rates, begin_date, end_date))
# print(form_link(id_rates, begin_date, end_date) == test_link)






"""
wb2 = load_workbook('Tue, 27 Sep 2022 07-06-04.xlsx')
print(wb2.sheetnames)

dest_filename = 'Tue, 27 Sep 2022 07-06-04.xlsx'
ws1 = wb.active
ws1.title = "range names"

for row in range(1, 40):
    ws1.append(range(600))

ws2 = wb.create_sheet(title="Pi")
ws2['F5'] = 3.14
ws3 = wb.create_sheet(title="Data")
for row in range(10, 20):
    for col in range(27, 54):
        _ = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))
print(ws3['AA10'].value)
wb.save(filename = dest_filename)
"""




